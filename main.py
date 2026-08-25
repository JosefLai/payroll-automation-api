from fastapi import FastAPI, File, UploadFile, Header, HTTPException
from fastapi.responses import StreamingResponse, JSONResponse
import pandas as pd
import numpy as np
import io
import os
from openpyxl.styles import PatternFill

app = FastAPI(title="Payroll Automation API")

# ----------------------------------------------------
# Simple shared-secret auth. This endpoint receives real payroll data (names, hours,
# wages), so it should not be callable by anyone who finds the Render URL. Set the
# PAYROLL_API_KEY environment variable in the Render dashboard, and configure Make.com's
# HTTP module to send the same value in an "X-API-Key" header. If PAYROLL_API_KEY is not
# set on the server, the check is skipped (local testing only - do not deploy this way).
# ----------------------------------------------------
EXPECTED_API_KEY = os.environ.get("PAYROLL_API_KEY")


def verify_api_key(x_api_key: str = Header(default=None)):
    if EXPECTED_API_KEY and x_api_key != EXPECTED_API_KEY:
        raise HTTPException(status_code=401, detail="Invalid or missing API key.")

# ----------------------------------------------------
# Rule 4: Employee-specific loaded rates
# ----------------------------------------------------
EMPLOYEE_RATES = {
    "DEFAULT": {"rate": 40.0, "loading": 18.0},          # Adj Rate = $58.00
    "ayman": {"rate": 50.0, "loading": 22.50},            # Adj Rate = $72.50
    "christian": {"rate": 27.0, "loading": 12.15},        # Adj Rate = $39.15
}

# ----------------------------------------------------
# Rule 1 & 2: Pruning rules (confirmed with Dan, Aug 4-5 email thread)
# ----------------------------------------------------
EXCLUDED_JOBCODES = ["general admin", "sick", "sick- unpaid", "management"]
# Vacation / Holiday are NOT dropped outright (per Dan's Aug 13 question + Josef's Aug 15
# reply) - they count toward the weekly 44h OT threshold using their actual QBO-recorded
# hours (a full day is typically 8h, but genuine partial days - e.g. 5.5h - are preserved
# as-is, not overridden to a flat 8). They contribute no OT themselves, and are dropped
# from the final billable output after they've done their job of shifting the weekly
# cumulative total, since they aren't job-costed. Per Dan's Aug 24 note, Stat holidays are
# always entered as "Holiday" (not "Vacation"), but both codes are handled identically here
# in case a genuine vacation day is ever recorded.
THRESHOLD_ONLY_JOBCODES = ["vacation", "holiday"]

# ASSUMPTION (open item - see email): these 3 employees are hard-coded exceptions because
# they log admin time under jobcodes shared with billable staff (e.g. "Quotes", "My Home
# Handyman"). Confirmed with Dan on Aug 4. Preferred long-term fix is asking them to stop
# using those jobcodes rather than growing this hard-coded list.
EXCLUDED_EMPLOYEES = ["mark biggins", "william lebherz", "jonathan kruger"]

# Penny-accurate minute table per Dan's Aug 5 request (60 values, 4 significant figures,
# rounded to the minute - not the second, per Dan's confirmation).
MINUTE_FRACTION = {m: round(m / 60, 4) for m in range(60)}

LUNCH_PAID_THRESHOLD_HR = 0.5  # Dan: first 30 min of a lunch break is paid, anything beyond is a flag


def get_employee_rate(fname, lname):
    full_name = f"{str(fname).strip()} {str(lname).strip()}".strip().lower()
    for name_key, rates in EMPLOYEE_RATES.items():
        if name_key != "DEFAULT" and name_key in full_name:
            return rates["rate"], rates["loading"]
    return EMPLOYEE_RATES["DEFAULT"]["rate"], EMPLOYEE_RATES["DEFAULT"]["loading"]


def flag_lunch_breaks(df):
    """Flag lunch break entries identified by jobcode_1 == 'Lunch break' (the literal
    QBO code Dan enters). Only flags the ones over the 30-minute paid threshold, since
    that's the case Dan needs to review. This is a REVIEW flag only - the automation
    does not change any hours based on it."""
    is_lunch = df["jobcode_1"].str.lower() == "lunch break"
    over_threshold = is_lunch & (df["hours"] > LUNCH_PAID_THRESHOLD_HR)
    df["Lunch_Flag"] = ""
    df.loc[over_threshold, "Lunch_Flag"] = "Lunch break >30 min - review"
    return df


def sequential_ot_split(df):
    """Alberta 8/44 OT split - CONFIRMED with Dan's Aug 10 email (with worked examples
    that this function reproduces exactly). This REPLACES the earlier proportional
    'shift_ratio' approach, which testing showed did not match Dan's actual method.

    The real rule, in Dan's own words: entries are processed in strict chronological
    order (earliest shift to latest) for each employee. A running total of hours worked
    TODAY resets at the start of each day; a running total of REGULAR hours worked THIS
    WEEK resets at the start of each ISO week (Mon-Sun). For each shift, in order:
      1. Daily 8h line: whatever portion of daily hours already worked pushes past 8.0
         within THIS shift becomes daily OT; nothing before or after this shift is
         touched by this step. Shifts entirely before the 8h line stay 100% Reg. Shifts
         entirely after it are 100% OT (from the day-threshold's perspective).
      2. Weekly 44h line: applied the same way to the day's resulting Reg hours, using
         the RUNNING WEEKLY REGULAR TOTAL (built from each day's own Reg result, not raw
         hours). Whichever shift's Reg hours cause the 44h line to be crossed converts
         only the portion beyond 44 to OT; everything chronologically after that stays
         100% OT for the rest of the week (see Dan's Sun example: 0 Reg, all OT, because
         the 44h Reg cap was already hit partway through Saturday).
    No proportional splitting anywhere. No special-casing for Lunch break/Downtime -
    testing confirmed this algorithm naturally reproduces Dan's actual treatment of both
    (validated: 1 of 78 Lunch break rows carries OT, 20 of 164 Downtime rows carry OT,
    both exact matches to Dan's real June working paper) WITHOUT needing to exempt them.

    Vacation/Holiday rows (is_threshold_only) are intentionally left in this same
    sequence - per Dan's Aug 13/Aug 15 confirmation, their actual QBO hours must count
    toward BOTH the daily 8h and weekly 44h running totals for the rest of that employee's
    week, exactly like a worked shift. They are dropped from the final output afterward.
    """
    df = df.sort_values(["full_name", "local_date", "local_start_time"]).reset_index(drop=True)
    df["year_week"] = df["local_date"].dt.strftime("%G-%V")

    reg_out = np.zeros(len(df))
    ot_out = np.zeros(len(df))

    for _, g in df.groupby("full_name", sort=False):
        daily_cum = 0.0
        weekly_cum_reg = 0.0
        cur_day = None
        cur_week = None
        for i in g.index:
            row = df.loc[i]
            if row["local_date"] != cur_day:
                daily_cum = 0.0
                cur_day = row["local_date"]
            if row["year_week"] != cur_week:
                weekly_cum_reg = 0.0
                cur_week = row["year_week"]

            d = row["hours"]

            # Step 1: daily 8h line
            if daily_cum >= 8.0:
                day_reg, day_ot = 0.0, d
            elif daily_cum + d <= 8.0:
                day_reg, day_ot = d, 0.0
            else:
                day_reg = 8.0 - daily_cum
                day_ot = d - day_reg
            daily_cum += d

            # Step 2: weekly 44h line, applied to this shift's day_reg portion
            if weekly_cum_reg >= 44.0:
                final_reg, final_ot = 0.0, day_ot + day_reg
            elif weekly_cum_reg + day_reg <= 44.0:
                final_reg, final_ot = day_reg, day_ot
            else:
                final_reg = 44.0 - weekly_cum_reg
                converted = day_reg - final_reg
                final_ot = day_ot + converted
            weekly_cum_reg += final_reg

            reg_out[i] = final_reg
            ot_out[i] = final_ot

    df["Reg"] = reg_out
    df["OT"] = ot_out
    return df


def calculate_alberta_ot(df):
    # ----------------------------------------------------
    # Data cleaning & column pre-processing
    # ----------------------------------------------------
    df["hours"] = pd.to_numeric(df["hours"], errors="coerce").fillna(0)
    df["full_name"] = df["fname"].astype(str).str.strip() + " " + df["lname"].astype(str).str.strip()
    df["full_name_clean"] = df["full_name"].str.lower().str.strip()

    # NOTE: fillna("") BEFORE astype(str) - otherwise genuinely blank cells (e.g. Lunch
    # break / Holiday rows with no class or service item) get rendered as the literal
    # text "nan" instead of staying blank. (Bug found from Dan's Aug 13 report, fixed
    # Aug 15.)
    for col in ["jobcode_1", "jobcode_2", "class", "service item"]:
        if col not in df.columns:
            df[col] = ""
        df[col] = df[col].fillna("").astype(str).str.strip()
        df[col] = df[col].replace(to_replace=r"(?i)^nan$", value="", regex=True)

    df["local_date"] = pd.to_datetime(df["local_date"])
    df["local_start_time"] = pd.to_datetime(df["local_start_time"])
    df["local_end_time"] = pd.to_datetime(df["local_end_time"])

    # ----------------------------------------------------
    # Rule 3: drop hours = 0 (QBO raw export column, confirmed with Dan Jul 24)
    # ----------------------------------------------------
    df = df[df["hours"] > 0].copy()

    # ----------------------------------------------------
    # Rule 1 & 2: pruning (jobcodes, admin class/service item, exception employees)
    # ----------------------------------------------------
    is_excluded_user = df["full_name_clean"].isin(EXCLUDED_EMPLOYEES)
    is_excluded_jobcode = df["jobcode_1"].str.lower().isin(EXCLUDED_JOBCODES)
    is_admin_class = (
        df["class"].str.lower().str.contains("admin", na=False)
        | df["service item"].str.lower().str.contains("admin", na=False)
    )
    df = df[~(is_excluded_user | is_excluded_jobcode | is_admin_class)].copy()

    if df.empty:
        return df

    # ----------------------------------------------------
    # Rule 1b: Vacation / Holiday - count toward the weekly 44h OT threshold using their
    # actual QBO-recorded hours (not forced to a flat 8h - genuine partial days, e.g.
    # 5.5h, are preserved as-is). They contribute no OT themselves and are removed from
    # the final billable output below, once they've done their job of shifting the
    # running daily/weekly totals for the rest of the employee's real shifts that day/week.
    # ----------------------------------------------------
    df["is_threshold_only"] = df["jobcode_1"].str.lower().isin(THRESHOLD_ONLY_JOBCODES)

    # ----------------------------------------------------
    # Lunch break >30 min flag (identified by jobcode_1)
    # ----------------------------------------------------
    df = flag_lunch_breaks(df)

    # ----------------------------------------------------
    # Alberta 8/44 OT split - sequential method confirmed with Dan (Aug 10 email).
    # Vacation/Holiday rows stay in this pass (see sequential_ot_split docstring) and are
    # dropped from the output afterward.
    # ----------------------------------------------------
    df = sequential_ot_split(df)

    # Flag shifts that straddle an OT threshold crossing (partial Reg + partial OT within
    # the SAME entry) - these are the highest-value rows to spot-check, since they're
    # exactly where the 8h or 44h line fell.
    df["Multi_Shift_Day"] = np.where((df["Reg"] > 0) & (df["OT"] > 0), "OT threshold crossed mid-shift - review", "")

    # ----------------------------------------------------
    # Keep full floating-point precision from Reg/OT through to the Hours/Minutes split.
    # Rounding Reg/OT to 2 decimals BEFORE deriving Hours/Minutes double-rounds and can
    # land on the wrong minute (e.g. 11.05 exact -> minute 3, but 11.04 rounded first ->
    # minute 2). Reg/OT/Adj Total below are rounded only for on-screen display;
    # Hours/Minutes/Dollars use the un-rounded value.
    # ----------------------------------------------------
    reg_raw = df["Reg"].copy()
    ot_raw = df["OT"].copy()
    adj_total_raw = reg_raw + ot_raw * 1.5

    df["Reg"] = reg_raw.round(2)
    df["OT"] = ot_raw.round(2)
    df["Adj Total"] = adj_total_raw.round(2)

    # 4. Jobber hours/minutes split (pink / purple columns)
    df["Hours"] = adj_total_raw.apply(lambda x: int(np.floor(x)))
    df["Minutes"] = adj_total_raw.apply(lambda x: int(round((x % 1) * 60)))
    rolled_over = df["Minutes"] == 60
    df.loc[rolled_over, "Hours"] += 1
    df.loc[rolled_over, "Minutes"] = 0

    # ----------------------------------------------------
    # Rate + penny-accurate Dollars (per Dan's Aug 5 request: 60-value, 4-sig-fig table)
    # ----------------------------------------------------
    rates = df.apply(lambda r: get_employee_rate(r["fname"], r["lname"]), axis=1)
    df["Rate"] = [r[0] for r in rates]
    df["Loading"] = [r[1] for r in rates]
    df["Adj Rate"] = df["Rate"] + df["Loading"]
    df["Dollars"] = ((df["Hours"] + df["Minutes"].map(MINUTE_FRACTION)) * df["Adj Rate"]).round(2)

    # Job / Cost Code - Phase 2 (not yet built). Left blank intentionally.
    df["Job"] = ""

    # ----------------------------------------------------
    # Rule 1b (cont.): Vacation/Holiday rows have now correctly consumed their share of
    # the daily/weekly OT thresholds above - drop them from the final billable output
    # since they are not job-costed.
    # ----------------------------------------------------
    df = df[~df["is_threshold_only"]].copy()

    output_cols = [
        "fname", "lname", "local_date", "local_day", "local_start_time", "local_end_time",
        "hours", "Reg", "OT", "Adj Total", "Hours", "Minutes",
        "Rate", "Loading", "Adj Rate", "Dollars", "Job",
        "jobcode_1", "jobcode_2", "class", "service item", "notes", "approved_status",
        "Lunch_Flag", "Multi_Shift_Day",
    ]
    for col in output_cols:
        if col not in df.columns:
            df[col] = ""
    df = df[output_cols]

    # Multi-level sort: 1. jobcode_1  2. local_date (old->new)  3. fname (A-Z)
    df = df.sort_values(by=["jobcode_1", "local_date", "fname"], ascending=[True, True, True]).reset_index(drop=True)

    return df


REQUIRED_COLUMNS = ["fname", "lname", "local_date", "local_start_time", "local_end_time", "hours", "jobcode_1"]


@app.get("/")
def health_check():
    return {"status": "ok", "message": "Payroll API is running successfully."}


@app.post("/process-timesheet")
async def process_timesheet(file: UploadFile = File(...), x_api_key: str = Header(default=None)):
    verify_api_key(x_api_key)

    contents = await file.read()
    try:
        df_raw = pd.read_csv(io.BytesIO(contents))
    except Exception:
        return JSONResponse(
            status_code=400,
            content={"error": "Could not read the file as a CSV. Please confirm it's the raw QBO timesheet export."},
        )

    missing = [c for c in REQUIRED_COLUMNS if c not in df_raw.columns]
    if missing:
        return JSONResponse(
            status_code=400,
            content={"error": f"The uploaded file is missing expected column(s): {', '.join(missing)}. "
                               f"This usually means it isn't the raw QBO export, or QBO changed its export format."},
        )

    try:
        df_processed = calculate_alberta_ot(df_raw)
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"error": f"Processing failed: {str(e)}. Please forward this file to Josef for review."},
        )

    if df_processed.empty:
        return JSONResponse(
            status_code=400,
            content={"error": "No billable rows remained after pruning. Please check the uploaded file."},
        )

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_processed.to_excel(writer, index=False, sheet_name="Processed Payroll")
        ws = writer.sheets["Processed Payroll"]

        hours_col_idx = df_processed.columns.get_loc("Hours") + 1
        minutes_col_idx = df_processed.columns.get_loc("Minutes") + 1
        lunch_col_idx = df_processed.columns.get_loc("Lunch_Flag") + 1
        multi_col_idx = df_processed.columns.get_loc("Multi_Shift_Day") + 1

        fill_hours = PatternFill(start_color="EAD1DC", end_color="EAD1DC", fill_type="solid")
        fill_minutes = PatternFill(start_color="D9D2E9", end_color="D9D2E9", fill_type="solid")
        fill_flag = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        for row in range(1, ws.max_row + 1):
            ws.cell(row=row, column=hours_col_idx).fill = fill_hours
            ws.cell(row=row, column=minutes_col_idx).fill = fill_minutes
            if row > 1:
                if ws.cell(row=row, column=lunch_col_idx).value:
                    ws.cell(row=row, column=lunch_col_idx).fill = fill_flag
                if ws.cell(row=row, column=multi_col_idx).value:
                    ws.cell(row=row, column=multi_col_idx).fill = fill_flag

    output.seek(0)

    filename = file.filename.replace(".csv", "_Processed.xlsx")
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
